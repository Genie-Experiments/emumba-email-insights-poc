import openpyxl
import re


def load_excel_file_cleaned(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    headers = [
        str(cell.value).strip() if cell is not None else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    headers = [header for header in headers if header]

    account_name_idx = headers.index("Account Name")
    owner_email_idx = headers.index("Account Owner Email")
    solution_eng_email_idx = headers.index("Solution Engineer Email")
    result = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = [str(cell).strip() if cell is not None else "" for cell in row]

        company = row_data[account_name_idx]
        owner_email = row_data[owner_email_idx]
        solution_eng_email = row_data[solution_eng_email_idx]

        if company and (owner_email or solution_eng_email):
            row_dict = {
                "target_customer": company,
                "target_emails": [owner_email, solution_eng_email],
            }
            result.append(row_dict)

    return result


def to_snake_case(s):
    return re.sub(r"\s+", "_", re.sub(r"[^\w\s]", "", s)).lower()
